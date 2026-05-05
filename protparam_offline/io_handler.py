"""
Input/output handling for protparam_offline.

Supported INPUT formats (auto-detected from file extension):
    .csv             Comma-separated values
    .tsv / .txt      Tab-separated values
    .xlsx / .xls     Excel workbook (uses first sheet by default)
    .fasta / .fa     FASTA-formatted protein sequences

Output is always a multi-sheet Excel (.xlsx) workbook.
"""

from __future__ import annotations
import os
from typing import List, Optional, Tuple

import pandas as pd

from .analysis import ProteinAnalyzer

# Common column names that might hold the sequence — checked case-insensitively
SEQUENCE_COLUMN_CANDIDATES = [
    "sequence", "seq", "protein_sequence", "aa_sequence",
    "peptide", "peptide_sequence", "protein", "amino_acid_sequence",
]
# Common ID column names
ID_COLUMN_CANDIDATES = [
    "protein_id", "id", "name", "uniprot_id", "accession",
    "header", "rank", "identifier",
]


# ---------- file reading ----------------------------------------------

def _read_fasta(path: str) -> pd.DataFrame:
    """Parse a FASTA file into a DataFrame with Protein_ID and sequence columns."""
    from Bio import SeqIO
    rows = []
    for record in SeqIO.parse(path, "fasta"):
        rows.append({
            "Protein_ID": record.id,
            "Description": record.description,
            "sequence": str(record.seq),
        })
    if not rows:
        raise ValueError(f"No sequences found in FASTA file: {path}")
    return pd.DataFrame(rows)


def read_input_file(path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Read an input file and return a pandas DataFrame.

    The format is detected from the file extension.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Input file not found: {path}")

    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(path)
    if ext in (".tsv", ".txt"):
        return pd.read_csv(path, sep="\t")
    if ext in (".xlsx", ".xls"):
        if sheet_name is None:
            return pd.read_excel(path)
        return pd.read_excel(path, sheet_name=sheet_name)
    if ext in (".fasta", ".fa", ".faa", ".fas"):
        return _read_fasta(path)

    raise ValueError(
        f"Unsupported file type '{ext}'. "
        f"Supported: .csv, .tsv, .txt, .xlsx, .xls, .fasta, .fa, .faa, .fas"
    )


# ---------- column detection ------------------------------------------

def detect_sequence_column(df: pd.DataFrame) -> Optional[str]:
    """Return the first column whose name matches a common sequence-column name."""
    lower_to_real = {c.lower().strip(): c for c in df.columns}
    for candidate in SEQUENCE_COLUMN_CANDIDATES:
        if candidate in lower_to_real:
            return lower_to_real[candidate]
    return None


def detect_id_column(df: pd.DataFrame) -> Optional[str]:
    """Return the first column whose name matches a common identifier-column name."""
    lower_to_real = {c.lower().strip(): c for c in df.columns}
    for candidate in ID_COLUMN_CANDIDATES:
        if candidate in lower_to_real:
            return lower_to_real[candidate]
    return None


# ---------- analysis pipeline -----------------------------------------

def analyze_dataframe(
    df: pd.DataFrame,
    sequence_column: str,
    id_column: Optional[str] = None,
    pH: float = 7.0,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Run the analyzer on every row of `df`. Returns (results_df, warnings).
    """
    if sequence_column not in df.columns:
        raise KeyError(
            f"Column '{sequence_column}' not found in input. "
            f"Available columns: {list(df.columns)}"
        )

    results = []
    warnings: List[str] = []

    for idx, row in df.iterrows():
        raw_seq = row[sequence_column]
        if id_column and id_column in df.columns:
            pid = str(row[id_column])
        else:
            pid = f"seq_{idx + 1}"

        try:
            analyzer = ProteinAnalyzer(raw_seq, protein_id=pid)
            if analyzer.length == 0:
                warnings.append(f"Row {idx + 1} ({pid}): empty/invalid sequence — skipped")
                continue
            results.append(analyzer.summary(pH=pH))
        except Exception as exc:
            warnings.append(f"Row {idx + 1} ({pid}): {exc}")

    return pd.DataFrame(results), warnings


# ---------- excel output ----------------------------------------------

def write_excel_output(
    output_path: str,
    original_df: pd.DataFrame,
    results_df: pd.DataFrame,
    warnings: Optional[List[str]] = None,
) -> None:
    """
    Write a multi-sheet Excel workbook:
        - Original_Data       (the input as it was read in)
        - Calculated_Results  (one row per successfully analyzed sequence)
        - Summary             (mean / min / max of each numeric column)
        - Warnings            (only if any warnings were produced)
    """
    # Make sure we can write to it; auto-append .xlsx if missing
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    # Build a numeric summary
    numeric_cols = results_df.select_dtypes(include="number").columns
    summary_df = (
        results_df[numeric_cols].describe().transpose().reset_index()
        .rename(columns={"index": "Metric"})
        if len(numeric_cols) and len(results_df)
        else pd.DataFrame({"Note": ["No numeric results to summarize."]})
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        original_df.to_excel(writer, sheet_name="Original_Data", index=False)
        results_df.to_excel(writer, sheet_name="Calculated_Results", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        if warnings:
            pd.DataFrame({"Warning": warnings}).to_excel(
                writer, sheet_name="Warnings", index=False
            )

    # Light formatting: bold header row, auto column widths
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        wb = load_workbook(output_path)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value is not None),
                    default=10,
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        wb.save(output_path)
    except Exception:
        # Formatting is cosmetic — never fail the run because of it
        pass
